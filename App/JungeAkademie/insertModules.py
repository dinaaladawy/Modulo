# -*- coding: utf-8 -*-
"""
Created on Mon Mar 13 14:11:11 2017

@author: Andrei
"""

import django, os, re, datetime
os.environ['DJANGO_SETTINGS_MODULE'] = "JungeAkademie.settings"
django.setup()
from openpyxl import load_workbook
from modulo.models import Category, Module, Language, Exam, CourseFormat, Location
#from openpyxl.worksheet.worksheet import Worksheet

moduleFile = "./modules.xlsx"

def getCategory(cell=None):
    if cell.value == None:
        return None
    else:
        try:
            return Category.objects.get(name=cell.value)
        except Category.DoesNotExist:
            print("Category", cell.value, "does not exist!!!")
            return None

def getLocation(cell=None):
    return Location.NOT_SPECIFIED

def getTime(cell=None):
    return datetime.time(hour=0, minute=0, second=0)

def getOrganisers(cell=None):
    if cell.value == None:
        return ""
    
    return cell.value.replace("_", "; ")

def getCredits(cell=None):
    if cell.value == None:
        return 0
    
    try:
        val = int(cell.value)
    except ValueError:
        # it was a string -> different number of credits depending on exam type
        # !!!for now take the max number of credits!!!
        s = cell.value
        l = [int(x) for x in filter(None, re.split("[\-/]+", s))]
        val = max(l)
    return val

def getLanguage(cell=None):
    val = []
    if cell.value == None:
        return [Language.NOT_SPECIFIED]
    
    if 'german' in cell.value.lower() or 'deutsch' in cell.value.lower():
        val.append(Language.GERMAN)
    if 'english' in cell.value.lower() or 'englisch' in cell.value.lower():
        val.append(Language.ENGLISH)
    if val == []:
        val.append(Language.OTHER)
    return val

def getExam(cell=None):
    val = []
    if cell.value == None:
        return [Exam.NOT_SPECIFIED], ''
    
    if ('schriftlich' in cell.value.lower() and 'prüfung' in cell.value.lower()) or \
           'klausur' in cell.value.lower() or \
           'exam' in cell.value.lower() or \
            ('written' in cell.value.lower() and 'exam' in cell.value.lower()):
        val.append(Exam.WRITTEN_EXAM)
    if 'präsentation' in cell.value.lower() or \
            ('mündlich' in cell.value.lower() and 'prüfung' in cell.value.lower()) or \
            'kolloquium' in cell.value.lower() or \
            'presentation' in cell.value.lower() or \
            'colloquium' in cell.value.lower() or \
            ('oral' in cell.value.lower() and 'exam' in cell.value.lower()):
        val.append(Exam.ORAL_EXAM)
    if 'referat' in cell.value.lower() or \
            'paper' in cell.value.lower() or\
            'essay' in cell.value.lower():
        val.append(Exam.PAPER)
    if val == []:
        val.append(Exam.OTHER)
    return val, cell.value

def getType(cell=None):
    val = []
    if cell.value == None:
        return [CourseFormat.NOT_SPECIFIED]
    
    if 'seminar' in cell.value.lower():
        val.append(CourseFormat.SEMINAR)
    if 'workshop' in cell.value.lower():
        val.append(CourseFormat.WORKSHOP)
    if 'kolloquium' in cell.value.lower() or  'colloquium' in cell.value.lower():
        val.append(CourseFormat.COLLOQUIUM)
    if 'modul' in cell.value.lower() or 'module' in cell.value.lower():
        val.append(CourseFormat.MODULE)
    if 'vorlesung' in cell.value.lower() or 'course' in cell.value.lower():
        val.append(CourseFormat.COURSE)
    if 'übung' in cell.value.lower() or 'exercice' in cell.value.lower():
        val.append(CourseFormat.EXERCICE)
    if 'exkursion' in cell.value.lower() or 'excursion' in cell.value.lower():
        val.append(CourseFormat.EXCURSION)
    if 'vortragsreihe' in cell.value.lower() or 'presentation' in cell.value.lower():
        val.append(CourseFormat.PRESENTATION)
    if val == []:
        val.append(CourseFormat.NOT_SPECIFIED)
    return val

def updateOrCreateModule(attr, row):
    #primary key???!!!
    #title, exam, credits, language, place
    
    defaults = {}
    categoryList = []
    for col in range(len(attr)):
        try:
            cell = row[col]
        except IndexError:
            cell = None
            
        if attr[col] == 'title':
            title=cell.value
            #defaults.update({attr[col]: cell.value})
        elif attr[col] == 'organisers':
            val = getOrganisers(cell)
            defaults.update({attr[col]: val})
        elif attr[col] == 'credits':
            #cell.value should be a string (check if string and then convert to int)
            val = getCredits(cell)
            defaults.update({attr[col]: val})
            #print("Number of credits of module ", title, ": ", val, sep='')
        elif attr[col] == 'exam_details':
            val, detail_val = getExam(cell)
            defaults.update({'exam': val[0], attr[col]: detail_val})
        elif attr[col] == 'type':
            val = getType(cell)
            defaults.update({attr[col]: val[0]})
        elif attr[col] == 'language':
            val = getLanguage(cell)
            defaults.update({attr[col]: val[0]})
        elif 'category' in attr[col]:
            categoryList.append(getCategory(cell))
        elif 'defaults' in attr[col]:
            #!!!
            #place and time are not yet in database...
            #!!!
            defaults.update({'place': getLocation(cell)})
            defaults.update({'time': getTime(cell)})
        else:
            #the rest of values are ok to just be copied
            if cell.value != None:
                defaults.update({attr[col]: cell.value})
    #defaults.update({'categories': []})
    
    try:
        old = Module.objects.get(title=title)
    except Module.DoesNotExist:
        old = Module()
        
    m, created = Module.objects.update_or_create(title=title, defaults=defaults)
    categoryList = list(filter(None, categoryList))
    #if m.categories.all() != categoryList:
        #print("Updated categories for module", m)
        #print(categoryList)
    try:
        m.categories.set(categoryList) #!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
    except django.core.exceptions.FieldDoesNotExist:
        print("Error on module", m)
    
    updated = (not created and old != m)
    if updated:
        print("Module ", m, " was updated.", sep='')
    return m, updated, created

def insertModules():
    #workbook = load_workbook(moduleFile, read_only=True)
    workbook = load_workbook(moduleFile)
    worksheet = workbook['Modules']
    nrRows = len(tuple(worksheet.rows))
    nrCols = len(tuple(worksheet.columns))
    attr = ['title', 'organisers', 'subtitle', 'description', 'goals', 'methods', 'sws', 'credits', 'type', 'exam_details', 'language', 'minParticipants', 'maxParticipants', 'category1', 'category2', 'category3', 'category4', 'category5', 'defaults']
    
    updates = 0
    creations = 0
    for row in worksheet.iter_rows(min_row=2, min_col=1, max_row=nrRows, max_col=nrCols):
        #this is a module
        m, updated, created = updateOrCreateModule(attr, row)
        updates += 1 if updated else 0
        creations += 1 if created else 0
        #print(m)
    
    print('Number of creations in database =', creations)
    print('Number of updates in database =', updates)
    print('Number of modules in database =', len(Module.objects.all()))
    workbook.close()
    
insertModules()
input("Press ENTER to exit program...")