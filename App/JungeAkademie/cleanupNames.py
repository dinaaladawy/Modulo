# -*- coding: utf-8 -*-
"""
Created on Wed Mar 22 20:11:47 2017

@author: Andrei
"""

import django, os
os.environ['DJANGO_SETTINGS_MODULE'] = "JungeAkademie.settings"
django.setup()
from modulo.models import Exam, Location, CourseFormat, Language, Personality
from modulo.models import Category, Interest, Module, TestPerson
from openpyxl import load_workbook

def cleanupString(s):
    #– -> -
    #{’ „ “ ” "} -> '
    try:
        s = s.replace("–", "-")
        #s = s.replace("„", "\"")
        s = s.replace("„", "'")
        #s = s.replace("“", "\"")
        s = s.replace("“", "'")
        #s = s.replace("”", "\"")
        s = s.replace("”", "'")
        s = s.replace("’", "'")
        s = s.replace('"', "'")
    except AttributeError:
        #print(s, "is not a string.")
        pass
    return s

def cleanupExcel():
    excelFile = "./database.xlsx"
    worksheetNames = ["interests", "categories", "modules", "exams", "languages", "courseFormats", "locations", "personalities"]
    workbook = load_workbook(excelFile)
    glob_counter = 0
    for name in worksheetNames:
        worksheet = workbook[name]
        nrRows = len(tuple(worksheet.rows))
        nrCols = len(tuple(worksheet.columns))
        sheet_counter = 0
        for row in worksheet.iter_rows(min_row=1, min_col=1, max_row=nrRows, max_col=nrCols):
            for cell in row:
                cleanup = cleanupString(cell.value)
                if cell.value != cleanup:
                    #print("Cleaned-up string:", cleanup, "\ncell.value: ", cell.value, "\n", sep='\n')
                    cell.value = cleanup
                    sheet_counter += 1
        if sheet_counter > 0:
            print("Cleaned-up %s entries in \'%s\' worksheet of the %s file" % (sheet_counter, name, excelFile))
            glob_counter += 1
    if glob_counter > 0:
        workbook.save(filename=excelFile)
	
def cleanupModules():
    mod_counter = 0
    cleanupAttributes = ['title', 'organisers', 'subtitle', 'description', 'goals', 'methods', 'exam_details']    
    for m in Module.objects.all():
        counter = 0
        for attr in cleanupAttributes:
            #old = m.title
            old = getattr(m, attr)
            #m.title = cleanupString(old)
            setattr(m, attr, cleanupString(old))
            if getattr(m, attr) != old:
                m.save()
                counter += 1
        if counter > 0:
            print("Cleaned-up %s entries in module %s" % (counter, m.title))
            mod_counter += 1
    print("Cleaned-up %s modules." % mod_counter)
        
def cleanupInterests():
    counter = 0
    for i in Interest.objects.all():
        old = i.name
        i.name = cleanupString(old)
        if i.name != old:
            i.save()
            counter += 1
    print("Cleaned-up %s interests." % counter)
        
def cleanupCategories():
    counter = 0
    for c in Category.objects.all():
        old = c.name
        c.name = cleanupString(old)
        if c.name != old:
            c.save()
            counter += 1
    print("Cleaned-up %s categories." % counter)
        
def cleanupExams():
    counter = 0
    for e in Exam.objects.all():
        old = e.exam_type
        e.name = cleanupString(old)
        if e.name != old:
            e.save()
            counter += 1
    print("Cleaned-up %s exams." % counter)

def cleanupCourseFormats():
    counter = 0
    for c in CourseFormat.objects.all():
        old = c.course_format
        c.course_format = cleanupString(old)
        if c.course_format != old:
            c.save()
            counter += 1
    print("Cleaned-up %s course formats." % counter)

def cleanupLanguages():
    counter = 0
    for l in Language.objects.all():
        old = l.language
        l.language = cleanupString(old)
        if l.language != old:
            l.save()
            counter += 1
    print("Cleaned-up %s languages." % counter)

def cleanupLocations():
    counter = 0
    for l in Location.objects.all():
        old = l.location
        l.location = cleanupString(old)
        if l.location != old:
            l.save()
            counter += 1
    print("Cleaned-up %s locations." % counter)

def cleanupPersonalities():
    counter = 0
    for p in Personality.objects.all():
        old = p.personality
        p.personality = cleanupString(old)
        if p.personality != old:
            p.save()
            counter += 1
    print("Cleaned-up %s personalities." % counter)

cleanupExcel()
cleanupExams()
cleanupModules()
cleanupInterests()
cleanupLanguages()
cleanupLocations()
cleanupCategories()
cleanupCourseFormats()
cleanupPersonalities()

if __name__ == '__main__':
    input("Press ENTER to exit programm...")