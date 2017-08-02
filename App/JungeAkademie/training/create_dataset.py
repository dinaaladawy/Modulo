# -*- coding: utf-8 -*-
"""
Created on 22 Jul 2017 at 13:55 

@author: Andrei
"""

from openpyxl import load_workbook
import django
import numpy as np
import json
import os

os.environ['DJANGO_SETTINGS_MODULE'] = "JungeAkademie.settings"
django.setup()
from modulo.models import Category, Interest, Module


def process_range(data_range):
    data = []
    for row in data_range:
        assert (0.0 <= float(row[0].value) <= 100.0)
        assert (len(row) == 2)
        if float(row[0].value) > 0.0:
            data.append((row[1].value, float(row[0].value)))
    return data


def get_input(interests):
    all_interests = [i.name for i in Interest.objects.all()]
    return [int(i in interests) for i in all_interests]


def get_inputs(interests_arg, num_samples):
    interests = [x[0] for x in interests_arg]
    inputs = [get_input(interests)]
    sample = 0
    nr_chosen_interests = len(interests_arg)
    while sample < num_samples - 1:
        random = np.random.uniform(0, 1, size=(nr_chosen_interests,))
        interests = []
        for index, i in enumerate(interests_arg):
            if random[index] < i[1]:
                interests.append(i[0])
        # ensure that min. 1 interest is selected
        if interests is []:
            continue
        inputs.append(get_input(interests))
        sample += 1
    return inputs


def get_label(categories):
    all_categories = [c.name for c in Category.objects.all()]
    return [int(c in categories) for c in all_categories]


def get_labels(modules_arg, num_samples):
    categories = [c.name for m in modules_arg for c in Module.objects.get(title=m[0]).categories.all()]
    labels = [get_label(categories)]
    nr_chosen_modules = len(modules_arg)
    for sample in range(num_samples - 1):
        random = np.random.uniform(0, 1, size=(nr_chosen_modules, ))
        categories = []
        for index, m in enumerate(modules_arg):
            if random[index] < m[1]:
                categories.extend([c.name for c in Module.objects.get(title=m[0]).categories.all()])
        labels.append(get_label(categories))
    return labels


def process_data(worksheet, num_categories, num_interests, num_modules, num_samples=1):
    interests = process_range(worksheet.iter_rows(min_row=2, min_col=1, max_row=num_interests, max_col=2))
    modules = process_range(worksheet.iter_rows(min_row=2, min_col=3, max_row=num_modules, max_col=4))
    data = {'inputs': get_inputs(interests, num_samples), 'labels': get_labels(modules, num_samples)}
    assert (len(data['inputs']) == len(data['labels']) == num_samples)
    assert (all([len(data['inputs'][s]) == num_interests] for s in range(num_samples)))
    assert (all([len(data['labels'][s]) == num_categories] for s in range(num_samples)))
    return data


def read_training_data(file, num_categories, num_interests, num_modules, num_samples=1):
    workbook = load_workbook(file)
    dataset = {'train': {'inputs': [], 'labels': []},
               'test': {'inputs': [], 'labels': []},
               'validation': {'inputs': [], 'labels': []}}
    for worksheet in workbook.worksheets:
        print(worksheet.title)
        try:
            personality_no = float(worksheet.title)
        except ValueError as e:
            print(e.args[0])
            print("Not reading data from {}".format(worksheet.title))
            continue
        nr_rows = len(tuple(worksheet.rows))
        nr_cols = len(tuple(worksheet.columns))
        print(nr_rows, nr_cols)

        data = process_data(worksheet, num_categories, num_interests, num_modules, num_samples)
        dataset['train']['inputs'].extend(data['inputs'])
        dataset['train']['labels'].extend(data['labels'])

    print(dataset)
    json.dump(dataset, open("dataset/dataset.json", "w"), sort_keys=True)


if __name__ == '__main__':
    import argparse

    parser = argparse.ArgumentParser(description='Train a model.')
    parser.add_argument("samples", type=int, default=1000)  # number of samples per person in database
    args = parser.parse_args()
    nr_samples = args.samples

    nr_categories = len(Category.objects.all())
    nr_interests = len(Interest.objects.all())
    nr_modules = len(Module.objects.all())
    print("Categories: {}".format(nr_categories))
    print("Interests: {}".format(nr_interests))
    print("Modules: {}".format(nr_modules))
    read_training_data("trainingdata.xlsx", num_interests=nr_interests, num_samples=nr_samples,
                       num_categories=nr_categories, num_modules=nr_modules)
