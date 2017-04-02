# -*- coding: utf-8 -*-
"""
Created on Mon Mar 13 14:11:11 2017

@author: Andrei
"""

import django, os
os.environ['DJANGO_SETTINGS_MODULE'] = "JungeAkademie.settings"
django.setup()
from modulo.models import Exam
from openpyxl import load_workbook

examsFile = "../database.xlsx"

def insertExams(databaseFile):
    workbook = load_workbook(databaseFile)
    worksheet = workbook['exams']
    nrRows = len(tuple(worksheet.rows))
    nrCols = len(tuple(worksheet.columns))
    
    add = 0
    for row in worksheet.iter_rows(min_row=2, min_col=1, max_row=nrRows, max_col=nrCols):
        cell = row[0]
        if cell.value is None:
            continue
        i, added = Exam.objects.get_or_create(exam_type=cell.value.strip())
        add += 1 if added else 0

    print("Number of added exams =", add)
    print('Number of exams in database =', len(Exam.objects.all()))
    #workbook.close()

if __name__ == "__main__":
    insertExams(examsFile)
    input("Press ENTER to exit program...")