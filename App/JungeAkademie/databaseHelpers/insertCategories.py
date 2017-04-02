# -*- coding: utf-8 -*-
"""
Created on Mon Mar 13 14:11:11 2017

@author: Andrei
"""

import django, os
os.environ['DJANGO_SETTINGS_MODULE'] = "JungeAkademie.settings"
django.setup()
from modulo.models import Category
from openpyxl import load_workbook

categoryFile = "./categories.xlsx"

def insertCategories(databaseFile):
    #workbook = load_workbook(categoryFile, read_only=True)
    workbook = load_workbook(databaseFile)
    worksheet = workbook['categories']
    nrRows = len(tuple(worksheet.rows))
    nrCols = len(tuple(worksheet.columns))
    
    add = 0
    for row in worksheet.iter_rows(min_row=2, min_col=1, max_row=nrRows, max_col=nrCols):
        #this is a category
        #assert(nrCols == 1)
        cell = row[0]
        if cell.value is None:
            continue
        c, added = Category.objects.get_or_create(name=cell.value.strip())
        add += 1 if added else 0
        #print(m)
    
    print("Number of added categories =", add)
    print('Number of categories in database =', len(Category.objects.all()))
    #workbook.close()

if __name__ == "__main__":
    insertCategories(categoryFile)
    input("Press ENTER to exit program...")