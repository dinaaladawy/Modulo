# -*- coding: utf-8 -*-
"""
Created on Mon Mar 13 14:11:11 2017

@author: Andrei
"""

import django, os
os.environ['DJANGO_SETTINGS_MODULE'] = "JungeAkademie.settings"
django.setup()
from modulo.models import TestPerson
from openpyxl import load_workbook

personalitiesFile = "../database.xlsx"

def insertPersonalities(databaseFile):
    workbook = load_workbook(databaseFile)
    worksheet = workbook['personalities']
    nrRows = len(tuple(worksheet.rows))
    nrCols = len(tuple(worksheet.columns))
    
    add = 0
    for row in worksheet.iter_rows(min_row=2, min_col=1, max_row=nrRows, max_col=nrCols):
        cell = row[0]
        if cell.value is None:
            continue
        #i, added = TestPerson.objects.get_or_create(personality=cell.value.strip())
        add += 1 if False else 0

    print("Number of added personalities =", add)
    print('Number of personalities in database =', len(TestPerson.objects.all()))
    #workbook.close()

if __name__ == "__main__":
    insertPersonalities(personalitiesFile)
    input("Press ENTER to exit program...")